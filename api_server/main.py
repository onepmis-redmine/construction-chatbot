from fastapi import FastAPI, Request, UploadFile, File, HTTPException, Query
from pydantic import BaseModel
import chromadb
import torch
from transformers import AutoTokenizer, AutoModel
import httpx
import json
from fastapi.middleware.cors import CORSMiddleware
import os
from dotenv import load_dotenv
import gc
from pathlib import Path
import pandas as pd
from fastapi.staticfiles import StaticFiles
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import asyncio
import time
from datetime import datetime
from utils.logger import get_logger
from utils.session_manager import SessionManager
import shutil
from fastapi.responses import JSONResponse, FileResponse
import sys
import subprocess
import csv
import io
import tempfile
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re
import numpy as np
import psutil

# 모듈별로 로거 생성
logger = get_logger(__name__)

# 프로젝트 루트 디렉토리 설정
ROOT_DIR = Path(__file__).parent.parent
VECTOR_DB_PATH = ROOT_DIR / "vector_db"
DOCS_DIR = ROOT_DIR / "docs"
ENHANCED_FAQ_PATH = DOCS_DIR / "enhanced_qa_pairs.xlsx"
FRONTEND_BUILD_DIR = ROOT_DIR / "frontend" / "build"
LOGS_DIR = ROOT_DIR / "logs"  # 로그 디렉토리 추가
QUESTIONS_DIR = ROOT_DIR / "questions"  # 질문 저장 디렉토리 추가

# 디렉토리 존재 확인 및 생성
VECTOR_DB_PATH.mkdir(parents=True, exist_ok=True)
DOCS_DIR.mkdir(parents=True, exist_ok=True)
LOGS_DIR.mkdir(parents=True, exist_ok=True)  # 로그 디렉토리 생성
QUESTIONS_DIR.mkdir(parents=True, exist_ok=True)  # 질문 저장 디렉토리 생성

# 질문 저장 파일 경로
QUESTIONS_FILE = QUESTIONS_DIR / "saved_questions.csv"

load_dotenv()  # .env 파일 읽기

# 환경 변수에서 API 키 불러오기
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")

# 모델 캐시 디렉토리 설정
cache_dir = ROOT_DIR / "model_cache"
cache_dir.mkdir(parents=True, exist_ok=True)

# FastAPI 앱 생성
app = FastAPI()

# CORS 설정 - localhost 우선
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",  # 로컬 개발 환경 우선
        "https://construction-chatbot-api.onrender.com",
        "https://construction-chatbot-frontend.onrender.com",  # 프론트엔드 배포 URL 추가
        "*",  # 개발 중에는 모든 출처 허용 (프로덕션에서는 제거)
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 개발 환경 감지
def is_development_request(request: Request) -> bool:
    """요청이 로컬 개발 환경에서 온 것인지 확인합니다."""
    origin = request.headers.get("origin", "")
    return origin.startswith("http://localhost:")

# ChromaDB 클라이언트 설정 - 전역 컬렉션 이름 상수 정의
COLLECTION_NAME = "construction_manuals"  # 원래 컬렉션 이름으로 되돌림

# ChromaDB 클라이언트 설정
chroma_client = chromadb.PersistentClient(path=str(VECTOR_DB_PATH))
collection = None  # 시작 시에는 None으로 설정, 초기화 과정에서 생성

def get_collection():
    """현재 컬렉션을 가져오거나 없으면 생성합니다."""
    global collection
    try:
        if collection is None:
            try:
                # 기존 컬렉션 가져오기 시도
                collection = chroma_client.get_collection(name=COLLECTION_NAME)
                logger.info(f"기존 컬렉션을 가져왔습니다: {COLLECTION_NAME}")
            except Exception as e:
                logger.warning(f"기존 컬렉션을 가져오는데 실패했습니다: {e}")
                # 컬렉션 생성
                collection = chroma_client.create_collection(name=COLLECTION_NAME)
                logger.info(f"새 컬렉션을 생성했습니다: {COLLECTION_NAME}")
        
        # 컬렉션 유효성 검사
        dummy_result = collection.count()
        logger.info(f"컬렉션 항목 수: {dummy_result}")
        return collection
    except Exception as e:
        logger.error(f"컬렉션 가져오기/생성 중 오류: {e}")
        # 마지막 시도: 모든 컬렉션 삭제 후 새로 생성
        try:
            chroma_client.delete_collection(name=COLLECTION_NAME)
            collection = chroma_client.create_collection(name=COLLECTION_NAME)
            logger.info(f"컬렉션을 강제로 재생성했습니다: {COLLECTION_NAME}")
            return collection
        except Exception as e2:
            logger.error(f"컬렉션 강제 재생성 중 오류: {e2}")
            raise e2

# FAQ 데이터 로드
faq_data = None

def load_enhanced_faq():
    """구조화된 FAQ 데이터를 로드합니다."""
    global faq_data
    try:
        if not ENHANCED_FAQ_PATH.exists():
            logger.error(f"Enhanced FAQ file not found at: {ENHANCED_FAQ_PATH}")
            return False
            
        faq_data = pd.read_excel(ENHANCED_FAQ_PATH)
        
        if faq_data.empty:
            logger.error("Enhanced FAQ file is empty")
            return False
        
        # 데이터 컬럼 확인 로깅
        logger.info(f"FAQ 데이터 컬럼: {list(faq_data.columns)}")
        logger.info(f"FAQ 데이터 첫 행: \n{faq_data.iloc[0]}")
        
        # 데이터 형식 검증
        for col in ['question_variations', 'structured_answer', 'keywords']:
            if col not in faq_data.columns:
                logger.error(f"Required column '{col}' not found in FAQ data")
                return False
        
        # JSON 필드 파싱 검증
        for idx, row in faq_data.iterrows():
            try:
                for field in ['question_variations', 'structured_answer', 'keywords']:
                    if isinstance(row[field], str):
                        json.loads(row[field])
                    else:
                        logger.warning(f"Row {idx}: {field} is not a string, converting to string")
                        faq_data.at[idx, field] = json.dumps(row[field], ensure_ascii=False)
            except Exception as e:
                logger.error(f"Error parsing JSON in row {idx}: {e}")
                return False
            
        logger.info(f"Enhanced FAQ data loaded successfully: {len(faq_data)} entries found")
        return True
    except Exception as e:
        logger.error(f"Error loading enhanced FAQ: {e}")
        faq_data = None
        return False

# 시작시 FAQ 데이터 로드
load_enhanced_faq()

# 모델 설정
model_name = "sentence-transformers/all-MiniLM-L6-v2"
tokenizer = None
model = None

def load_model():
    global tokenizer, model
    if tokenizer is None:
        tokenizer = AutoTokenizer.from_pretrained(model_name, local_files_only=False)
    if model is None:
        model = AutoModel.from_pretrained(
            model_name,
            local_files_only=False
        )
        model.eval()  # 평가 모드로 설정
    return tokenizer, model

def unload_model():
    global tokenizer, model
    del tokenizer
    del model
    tokenizer = None
    model = None
    torch.cuda.empty_cache() if torch.cuda.is_available() else gc.collect()

# Mean Pooling 함수
def mean_pooling(model_output, attention_mask):
    token_embeddings = model_output[0]
    input_mask_expanded = attention_mask.unsqueeze(-1).expand(token_embeddings.size()).float()
    return torch.sum(token_embeddings * input_mask_expanded, 1) / torch.clamp(input_mask_expanded.sum(1), min=1e-9)

# 텍스트를 임베딩으로 변환하는 함수
def get_embeddings(texts, batch_size=32):
    """텍스트를 임베딩으로 변환합니다. 배치 처리를 지원하며 메모리 효율성을 위해 최대 배치 크기를 제한합니다."""
    if isinstance(texts, str):
        texts = [texts]
    
    # 빈 입력 확인
    if not texts:
        logger.warning("get_embeddings: 입력 텍스트가 비어 있습니다.")
        return []
        
    # 로드된 텍스트 길이 로깅
    logger.info(f"get_embeddings: 처리할 텍스트 {len(texts)}개")
    
    # 배치가 너무 크면 분할하여 재귀적으로 처리
    if len(texts) > batch_size:
        logger.info(f"배치 크기({len(texts)})가 제한({batch_size})을 초과하여 분할 처리합니다.")
        result = []
        for i in range(0, len(texts), batch_size):
            batch_texts = texts[i:i + batch_size]
            batch_embeddings = get_embeddings(batch_texts)
            result.extend(batch_embeddings)
        return result
    
    try:
        # 모델 로드
        tokenizer, model = load_model()
        
        # 인코딩
        encoded_input = tokenizer(
            texts,
            padding=True,
            truncation=True,
            max_length=512,
            return_tensors='pt'
        )
        
        # 임베딩 생성
        with torch.no_grad():
            model_output = model(**encoded_input)
        
        sentence_embeddings = mean_pooling(model_output, encoded_input['attention_mask'])
        sentence_embeddings = torch.nn.functional.normalize(sentence_embeddings, p=2, dim=1)
        
        # 텐서를 리스트로 변환
        embeddings_list = sentence_embeddings.tolist()
        
        # 메모리 정리
        unload_model()
        
        # 임베딩 형식 확인 및 로깅
        if embeddings_list and len(embeddings_list) > 0:
            logger.info(f"생성된 임베딩 형식: 차원 수 = {len(embeddings_list)}, 첫 임베딩 길이 = {len(embeddings_list[0])}")
            
            # 임베딩이 3차원 리스트인 경우 2차원으로 변환
            if isinstance(embeddings_list[0], list) and isinstance(embeddings_list[0][0], list):
                logger.info("3차원 임베딩을 2차원으로 변환합니다.")
                embeddings_list = [emb[0] for emb in embeddings_list]
            
            # 단일 임베딩인 경우 리스트로 변환
            elif not isinstance(embeddings_list[0], list):
                embeddings_list = [embeddings_list]
                logger.info("단일 임베딩을 리스트로 변환했습니다.")
            
            # 최종 형식 검증
            if not all(isinstance(emb, list) and all(isinstance(x, (int, float)) for x in emb) for emb in embeddings_list):
                logger.error("임베딩 형식이 올바르지 않습니다.")
                return [[0.0] * 384 for _ in range(len(texts))]
        
        return embeddings_list
    except Exception as e:
        logger.error(f"임베딩 생성 중 오류 발생: {e}")
        import traceback
        logger.error(f"상세 오류: {traceback.format_exc()}")
        
        # 오류 발생 시 빈 임베딩 반환 (적절한 차원의 0 벡터)
        return [[0.0] * 384 for _ in range(len(texts))]

# 세션 매니저 초기화
session_manager = SessionManager(storage_dir=str(ROOT_DIR / "sessions"))

# 질문 저장 함수
def save_question(query: str, answer: str = None, similarity: float = None, session_id: str = None):
    """사용자 질문과 답변을 CSV 파일에 저장합니다."""
    try:
        # 파일이 존재하지 않으면 헤더와 함께 생성
        file_exists = QUESTIONS_FILE.exists()
        
        # 현재 시간을 원하는 형식으로 포맷팅 (YYYY-MM-DD AM/PM H:MM)
        now = datetime.now()
        formatted_time = now.strftime("%Y-%m-%d %p %I:%M").replace("AM", "오전").replace("PM", "오후")
        
        # BOM을 사용하여 UTF-8로 저장 (Excel에서 한글 인코딩 문제 해결)
        with open(QUESTIONS_FILE, mode='a', newline='', encoding='utf-8-sig') as file:
            writer = csv.writer(file)
            
            # 파일이 새로 생성된 경우 헤더 추가
            if not file_exists:
                writer.writerow(['timestamp', 'session_id', 'query', 'answer', 'similarity'])
            
            # 질문과 답변 저장
            writer.writerow([
                formatted_time,
                session_id or 'unknown',
                query,
                answer or '',
                f"{similarity:.4f}" if similarity is not None else ''
            ])
        
        logger.info(f"질문과 답변 저장 완료: {query}")
        return True
    except Exception as e:
        logger.error(f"질문 저장 중 오류 발생: {e}")
        import traceback
        logger.error(f"상세 오류: {traceback.format_exc()}")
        return False

# 질문 형식 정의
class Question(BaseModel):
    query: str
    session_id: str = None

# OpenRouter 호출 함수
def call_openrouter(prompt: str, conversation_history=None) -> str:
    headers = {
        "Authorization": OPENROUTER_API_KEY,
        "Content-Type": "application/json"
    }
    
    # API 키가 없는 경우 기본 메시지 반환
    if not OPENROUTER_API_KEY or OPENROUTER_API_KEY == "your_openrouter_api_key_here":
        logger.error("OpenRouter API 키가 설정되지 않았습니다.")
        return "죄송합니다. OpenRouter API 키가 설정되지 않아 질문에 대한 답변을 제공할 수 없습니다. 관리자에게 문의해주세요."
    
    messages = []
    
    # 시스템 메시지 추가 (존댓말 지시)
    system_message = {
        "role": "system", 
        "content": "당신은 건설 정보 시스템에 대한 전문 지식을 갖춘 챗봇입니다. 항상 정중하고 공손한 존댓말로 답변해주세요."
    }
    messages.append(system_message)
    
    # 대화 기록이 있으면 추가
    if conversation_history:
        for msg in conversation_history:
            messages.append({
                "role": msg["role"],
                "content": msg["content"]
            })
    
    # 현재 사용자 메시지 추가
    messages.append({"role": "user", "content": prompt})
    
    # 대화 기록이 없는 경우에도 시스템 메시지는 유지
    if len(messages) <= 2:  # 시스템 메시지 + 현재 메시지만 있는 경우
        messages = [system_message, {"role": "user", "content": prompt}]
    
    payload = {
        "model": "openai/gpt-3.5-turbo",
        "messages": messages
    }
    
    try:
        logger.info(f"OpenRouter API 호출: {prompt[:50]}...")
        
        response = httpx.post(
            url="https://openrouter.ai/api/v1/chat/completions",
            headers=headers,
            json=payload,
            timeout=60.0,
            verify=False  # 테스트 환경에서만 사용
        )
        
        # 응답 상태 코드 로깅
        logger.info(f"OpenRouter API 응답 상태 코드: {response.status_code}")
        
        # 응답 내용 디버그를 위해 로깅 (개인정보는 제외)
        logger.info(f"OpenRouter API 응답 헤더: {dict(response.headers)}")
        
        if response.status_code != 200:
            error_msg = f"OpenRouter API 오류: 상태 코드 {response.status_code}"
            try:
                error_data = response.json()
                error_msg += f", 상세 내용: {error_data}"
            except:
                error_msg += f", 응답: {response.text[:200]}"
            
            logger.error(error_msg)
            return f"죄송합니다. API 서버에 문제가 발생했습니다. 관리자에게 문의해주세요. (오류: {response.status_code})"
        
        data = response.json()
        
        # 응답 데이터 구조 확인
        if "choices" not in data or len(data["choices"]) == 0:
            logger.error(f"OpenRouter API 응답에 choices 필드가 없습니다: {data}")
            return "죄송합니다. API 응답이 올바르지 않습니다. 관리자에게 문의해주세요."
        
        if "message" not in data["choices"][0]:
            logger.error(f"OpenRouter API 응답에 message 필드가 없습니다: {data['choices'][0]}")
            return "죄송합니다. API 응답이 올바르지 않습니다. 관리자에게 문의해주세요."
        
        return data["choices"][0]["message"]["content"]
    except httpx.HTTPStatusError as e:
        logger.error(f"HTTP 오류: {e.response.status_code} - {e.response.text}")
        return f"죄송합니다. API 서버 응답 오류가 발생했습니다. (상태 코드: {e.response.status_code})"
    except httpx.RequestError as e:
        logger.error(f"요청 오류: {e}")
        return "죄송합니다. 서버 연결에 문제가 발생했습니다. 잠시 후 다시 시도해주세요."
    except Exception as e:
        logger.error(f"OpenRouter API 호출 중 오류 발생: {e}")
        import traceback
        logger.error(f"상세 오류: {traceback.format_exc()}")
        return f"죄송합니다. 질문 처리 중 오류가 발생했습니다. 잠시 후 다시 시도해주세요. (오류: {str(e)[:100]})"

def find_faq_match(query: str, threshold: float = 0.75):
    """구조화된 FAQ에서 가장 적절한 답변을 찾습니다."""
    
    if faq_data is None or faq_data.empty:
        logger.error("FAQ data is not loaded")
        return None, 0.0
    
    try:
        # 쿼리 전처리
        query = query.lower().strip()
        if query.endswith('?') or query.endswith('.'): 
            query = query[:-1]
        
        start_time = time.time()
        
        # 쿼리 임베딩 생성
        query_embedding = get_embeddings(query)
        
        # 단일 임베딩인 경우 리스트로 변환 (배치 처리 함수 변경으로 인한 대응)
        if isinstance(query_embedding, list) and not isinstance(query_embedding[0], list):
            query_embedding = [query_embedding]
            
        # 만약 임베딩 결과가 빈 리스트이거나 None이면 기본값 사용
        if not query_embedding or len(query_embedding) == 0:
            logger.error("임베딩 생성 실패, 기본값 사용")
            # 384차원 0 벡터 사용
            query_embedding = [[0.0] * 384]
        
        # 컬렉션 가져오기
        current_collection = get_collection()
        
        logger.info(f"벡터 검색 시작: {query}")
        
        # 기존 방식으로 돌아가기: 모든 FAQ 항목을 직접 비교
        best_match = None
        highest_similarity = -1
        
        # 쿼리에 포함된 단어 목록
        query_words = set(query.split())
        
        # 각 FAQ 항목과 비교
        for idx, row in faq_data.iterrows():
            try:
                variations = json.loads(row['question_variations']) if isinstance(row['question_variations'], str) else row['question_variations']
                original_question = row.get('original_question', '')
                
                # keywords 필드 활용
                keywords = []
                if 'keywords' in row:
                    try:
                        keywords_data = row['keywords']
                        if isinstance(keywords_data, str):
                            keywords = json.loads(keywords_data)
                        else:
                            keywords = keywords_data
                    except Exception as e:
                        logger.error(f"키워드 파싱 오류: {e}")
                
                logger.info(f"FAQ 항목 {idx + 1}/{len(faq_data)}: 원본 질문: '{original_question}', 키워드: {keywords}")
                
                # 키워드 매칭 점수 계산
                keyword_bonus = 0.0
                matching_keywords = []
                
                for keyword in keywords:
                    # 정확한 키워드 매칭 확인 (부분 매칭이 아닌 완전 일치)
                    # 1. 쿼리에 키워드가 정확히 포함되는지 확인 (단어 경계 고려)
                    exact_match = False
                    
                    # 키워드를 소문자로 변환하고 공백 처리
                    keyword_lower = keyword.lower().strip()
                    # 띄어쓰기를 제거한 버전도 생성
                    keyword_no_space = keyword_lower.replace(" ", "")
                    
                    # 쿼리에서 띄어쓰기 제거한 버전
                    query_no_space = query.lower().replace(" ", "")
                    # 쿼리 단어를 모두 붙인 버전
                    query_words_joined = "".join(query_words)
                    
                    # 방법 1: 쿼리에서 키워드를 단어 단위로 찾기
                    if keyword_lower in query_words:
                        exact_match = True
                        logger.info(f"단어 단위 매치: '{keyword}'")
                    
                    # 방법 2: 쿼리에서 키워드를 단어 경계를 고려하여 찾기
                    elif re.search(r'\b' + re.escape(keyword_lower) + r'\b', query.lower()):
                        exact_match = True
                        logger.info(f"단어 경계 매치: '{keyword}'")
                    
                    # 방법 3: 쿼리가 기본적으로 짧은 경우 정확한 키워드만 체크
                    elif len(query_words) <= 3 and query.lower() == keyword_lower:
                        exact_match = True
                        logger.info(f"전체 텍스트 매치: '{keyword}'")
                    
                    # 방법 4: 띄어쓰기를 무시하고 매칭
                    elif keyword_no_space == query_no_space:
                        exact_match = True
                        logger.info(f"띄어쓰기 무시 전체 매치: '{keyword}' vs '{query}'")
                    
                    # 방법 5: 띄어쓰기를 무시한 부분 문자열 매칭 (키워드가 꽤 길 경우에만)
                    elif len(keyword_no_space) >= 4 and keyword_no_space in query_no_space:
                        # 키워드가 충분히 길 경우에만 부분 문자열 매칭 허용
                        exact_match = True
                        logger.info(f"띄어쓰기 무시 부분 매치: '{keyword}' in '{query}'")
                    
                    if exact_match:
                        matching_keywords.append(keyword)
                        keyword_bonus += 0  # 키워드당 0.3점 보너스
                        logger.info(f"정확한 키워드 매치: '{keyword}'")
                
                if matching_keywords:
                    logger.info(f"키워드 매치 발견: {matching_keywords}, 보너스: {keyword_bonus}")
                
                for q in variations:
                    # 질문 전처리
                    q = q.lower().strip()
                    if q.endswith('?') or q.endswith('.'): 
                        q = q[:-1]
                        
                    # 직접 코사인 유사도 계산
                    try:
                        # q에 대한 임베딩 조회
                        q_embedding = get_embeddings(q)
                        
                        # 단일 임베딩인 경우 리스트로 변환
                        if isinstance(q_embedding, list) and not isinstance(q_embedding[0], list):
                            q_embedding = [q_embedding]
                        
                        # 코사인 유사도 계산
                        similarity = torch.cosine_similarity(
                            torch.tensor(query_embedding[0]),
                            torch.tensor(q_embedding[0]),
                            dim=0
                        ).item()
                        
                        # 키워드 보너스 적용
                        adjusted_similarity = similarity + keyword_bonus
                        
                        logger.info(f"질문: '{q}', 기본 유사도: {similarity:.4f}, 키워드 보너스: {keyword_bonus:.2f}, 최종: {adjusted_similarity:.4f}")
                        
                        if adjusted_similarity > highest_similarity:
                            highest_similarity = adjusted_similarity
                            best_match = row
                            logger.info(f"새 최고 매치: '{original_question}', 유사도: {adjusted_similarity:.4f}, 매칭 키워드: {matching_keywords}")
                    except Exception as e:
                        logger.error(f"유사도 계산 중 오류: {e}")
                        continue
            except Exception as e:
                logger.error(f"변형 질문 처리 중 오류: {e}")
                continue
        
        end_time = time.time()
        logger.info(f"검색 완료: 소요 시간 = {end_time - start_time:.2f}초, 최고 유사도 = {highest_similarity:.4f}")
        
        # 최종 유사도가 키워드 보너스 때문에 임계값을 넘었을 수 있으므로, 
        # 원래 임계값보다 크거나 같은지 확인
        if highest_similarity >= threshold and best_match is not None:
            return best_match, highest_similarity
        return None, highest_similarity if highest_similarity > -1 else 0.0
    except Exception as e:
        logger.error(f"Error in find_faq_match: {e}")
        import traceback
        logger.error(f"상세 오류: {traceback.format_exc()}")
        return None, 0.0

@app.post("/ask")
async def ask_question(q: Question, request: Request):
    """질문을 처리하고 답변을 반환합니다."""
    try:
        logger.info(f"Received question: {q.query}, session_id: {q.session_id}")
        start_time = time.time()
        
        # 개발 환경 여부 확인
        is_dev = is_development_request(request)
        
        # 세션 처리
        if not q.session_id:
            session = session_manager.create_session()
            q.session_id = session.session_id
            logger.info(f"Created new session: {q.session_id}")
        else:
            # 세션이 존재하는지 확인
            session = session_manager.get_session(q.session_id)
            if not session:
                # 존재하지 않는 세션 ID인 경우 새 세션 생성
                session = session_manager.create_session()
                q.session_id = session.session_id
                logger.info(f"Invalid session ID provided, created new: {q.session_id}")
        
        # 사용자 메시지 저장
        session_manager.add_message(q.session_id, "user", q.query)
        
        # FAQ 매칭 시도
        try:
            faq_match_start = time.time()
            faq_match, similarity = find_faq_match(q.query)
            faq_match_time = time.time() - faq_match_start
            
            logger.info(f"FAQ 매칭 소요 시간: {faq_match_time:.2f}초")
            
            if faq_match is not None:
                logger.info(f"FAQ match found for query: {q.query}")
                
                # faq_match 객체의 구조 확인
                logger.info(f"FAQ match columns: {list(faq_match.index)}")
                
                # 응답 데이터 생성
                response_data = {
                    "session_id": q.session_id,
                    "sources": [faq_match.get("source", "FAQ")],
                    "is_dev": is_dev,
                    "type": "faq",
                    "debug_info": {
                        "source": "faq",
                        "confidence": similarity,
                        "match_time": f"{faq_match_time:.2f}초"
                    }
                }
                
                # 원본 질문 추가
                if 'original_question' in faq_match:
                    response_data["original_question"] = faq_match["original_question"]
                
                # 답변 데이터 추가
                if 'answer' in faq_match:
                    # 단순 텍스트 답변
                    formatted_answer = faq_match["answer"]
                    response_data["answer"] = formatted_answer
                    session_manager.add_message(q.session_id, "assistant", formatted_answer)
                    
                    # 질문과 답변 저장
                    save_question(q.query, formatted_answer, similarity, q.session_id)
                    
                elif 'structured_answer' in faq_match:
                    # 구조화된 답변 - JSON 객체 그대로 전달
                    structured_answer = faq_match["structured_answer"]
                    if isinstance(structured_answer, str):
                        structured_answer = json.loads(structured_answer)
                    
                    # 구조화된 데이터 전달
                    response_data["structured_answer"] = structured_answer
                    
                    # 구조화된 답변을 읽기 쉬운 텍스트로 변환하여 세션에 저장
                    readable_text = ""
                    
                    # 원본 질문 추가
                    if "original_question" in faq_match:
                        readable_text += f"질문: {faq_match['original_question']}\n\n"
                    
                    # 기본 규칙 추가
                    if "basic_rules" in structured_answer and structured_answer["basic_rules"]:
                        readable_text += "• 기본 규칙:\n"
                        for rule in structured_answer["basic_rules"]:
                            readable_text += f"  - {rule}\n"
                        readable_text += "\n"
                    
                    # 예시 추가
                    if "examples" in structured_answer and structured_answer["examples"]:
                        readable_text += "• 예시:\n"
                        for example in structured_answer["examples"]:
                            if isinstance(example, dict):
                                # 딕셔너리 형태의 예시는 시나리오와 결과로 구분하여 처리
                                for key, value in example.items():
                                    if key.lower().startswith('scenario'):
                                        readable_text += f"  📌 {value}\n"
                                    elif key.lower().startswith('result'):
                                        readable_text += f"      ➡️ {value}\n"
                                    else:
                                        readable_text += f"      • {key}: {value}\n"
                            else:
                                readable_text += f"  - {example}\n"
                        readable_text += "\n"
                    
                    # 주의사항 추가
                    if "cautions" in structured_answer and structured_answer["cautions"]:
                        readable_text += "• 주의사항:\n"
                        for caution in structured_answer["cautions"]:
                            readable_text += f"  - {caution}\n"
                        readable_text += "\n"
                    
                    # 세션에 저장
                    session_manager.add_message(q.session_id, "assistant", readable_text)
                    
                    # 구조화된 답변을 문자열로 변환하여 저장
                    answer_text = json.dumps(structured_answer, ensure_ascii=False)
                    save_question(q.query, answer_text, similarity, q.session_id)
                    
                else:
                    # 답변 없음
                    formatted_answer = "죄송합니다. 이 질문에 대한 답변을 찾을 수 없습니다."
                    response_data["answer"] = formatted_answer
                    session_manager.add_message(q.session_id, "assistant", formatted_answer)
                    
                    # 질문과 답변 저장
                    save_question(q.query, formatted_answer, similarity, q.session_id)
                    
                    logger.error(f"FAQ match found but no answer field: {faq_match}")
                
                # 총 처리 시간 측정
                total_time = time.time() - start_time
                response_data["debug_info"]["total_time"] = f"{total_time:.2f}초"
                
                return response_data
            
            # OpenRouter API 호출
            openrouter_start = time.time()
            logger.info(f"No FAQ match found, calling OpenRouter for query: {q.query}")
            answer = call_openrouter(q.query, session_manager.get_messages(q.session_id))
            openrouter_time = time.time() - openrouter_start
            session_manager.add_message(q.session_id, "assistant", answer)
            
            # 질문과 답변 저장 (OpenRouter의 경우 similarity는 0으로 저장)
            save_question(q.query, answer, 0.0, q.session_id)
            
            # 총 처리 시간 측정
            total_time = time.time() - start_time
            
            return {
                "answer": answer,
                "session_id": q.session_id,
                "sources": [],
                "type": "openrouter",
                "is_dev": is_dev,
                "debug_info": {
                    "source": "openrouter",
                    "similarity": 0.0,
                    "faq_match_time": f"{faq_match_time:.2f}초",
                    "openrouter_time": f"{openrouter_time:.2f}초",
                    "total_time": f"{total_time:.2f}초"
                }
            }
        except Exception as e:
            logger.error(f"FAQ 매칭 중 오류: {e}")
            # OpenRouter로 폴백
            openrouter_start = time.time()
            logger.info(f"FAQ 매칭 실패, OpenRouter 호출: {q.query}")
            answer = call_openrouter(q.query, session_manager.get_messages(q.session_id))
            openrouter_time = time.time() - openrouter_start
            session_manager.add_message(q.session_id, "assistant", answer)
            
            # 질문과 답변 저장 (OpenRouter의 경우 similarity는 0으로 저장)
            save_question(q.query, answer, 0.0, q.session_id)
            
            # 총 처리 시간 측정
            total_time = time.time() - start_time
            
            return {
                "answer": answer,
                "session_id": q.session_id,
                "sources": [],
                "type": "openrouter",
                "is_dev": is_dev,
                "debug_info": {
                    "source": "openrouter",
                    "error": str(e),
                    "faq_match_time": f"{faq_match_time:.2f}초",
                    "openrouter_time": f"{openrouter_time:.2f}초",
                    "total_time": f"{total_time:.2f}초"
                }
            }
    except Exception as e:
        logger.error(f"Error processing question: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/session/{session_id}")
async def get_session(session_id: str):
    """세션의 대화 기록을 가져옵니다."""
    messages = session_manager.get_messages(session_id)
    if not messages:
        raise HTTPException(status_code=404, detail="Session not found")
    return {"messages": messages}

@app.post("/session/cleanup")
async def cleanup_sessions():
    """오래된 세션을 정리합니다."""
    session_manager.cleanup_old_sessions()
    return {"status": "success", "message": "Sessions cleaned up"}

@app.get("/health")
async def health_check():
    return {"status": "ok"}

@app.get("/reload-faq")
async def reload_faq():
    """FAQ 데이터를 재로드합니다."""
    load_enhanced_faq()
    return {"status": "FAQ reloaded"}

# 여기에서 initialize_vector_db 함수 수정
async def initialize_vector_db(force_rebuild=False):
    """구조화된 FAQ 데이터를 벡터 데이터베이스에 임베딩합니다."""
    global collection
    try:
        # 기존 컬렉션이 있는지 확인
        existing_collection = False
        try:
            test_collection = chroma_client.get_collection(name=COLLECTION_NAME)
            collection_count = test_collection.count()
            if collection_count > 0:
                existing_collection = True
                logger.info(f"기존 컬렉션이 존재함: {COLLECTION_NAME}, 항목 수: {collection_count}")
        except Exception as e:
            logger.warning(f"기존 컬렉션 확인 중 오류 발생: {e}")
            existing_collection = False
        
        # 기존 컬렉션이 있고 강제 재구축이 아니면 그대로 사용
        if existing_collection and not force_rebuild:
            collection = test_collection
            logger.info(f"기존 컬렉션을 재사용합니다: {COLLECTION_NAME}")
            return {"success": True, "message": f"기존 컬렉션 재사용 ({collection_count}개 항목)"}
        
        # 아래는 기존 임베딩 생성 코드 (새로 생성해야 하는 경우)
        if not load_enhanced_faq():
            logger.error("구조화된 FAQ 데이터를 로드할 수 없습니다.")
            return {"success": False, "message": "구조화된 FAQ 데이터를 로드할 수 없습니다."}
        
        if faq_data is None or faq_data.empty:
            logger.error("FAQ 데이터가 비어 있습니다.")
            return {"success": False, "message": "FAQ 데이터가 비어 있습니다."}
        
        logger.info("새 벡터 데이터베이스를 생성합니다...")
        
        # 기존 컬렉션 삭제 후 재생성
        try:
            chroma_client.delete_collection(name=COLLECTION_NAME)
            logger.info(f"기존 컬렉션 삭제 완료: {COLLECTION_NAME}")
        except Exception as e:
            logger.warning(f"컬렉션 삭제 중 오류 발생 (무시됨): {e}")
            
        # 새 컬렉션 생성
        collection = chroma_client.create_collection(name=COLLECTION_NAME)
        logger.info(f"새 컬렉션이 성공적으로 생성됨: {COLLECTION_NAME}")
        
        texts = []
        metadatas = []
        ids = []
        
        # 각 FAQ 항목에 대해 임베딩 생성
        total_items = len(faq_data)
        logger.info(f"총 처리할 FAQ 항목: {total_items}개")
        
        for idx, row in faq_data.iterrows():
            # 진행률 표시
            progress_percent = (idx + 1) / total_items * 100
            logger.info(f"FAQ 임베딩 진행률: {progress_percent:.1f}% ({idx + 1}/{total_items})")
            
            # JSON 문자열을 파싱
            question_variations = json.loads(row['question_variations']) if isinstance(row['question_variations'], str) else row['question_variations']
            structured_answer = json.loads(row['structured_answer']) if isinstance(row['structured_answer'], str) else row['structured_answer']
            keywords = json.loads(row['keywords']) if isinstance(row['keywords'], str) else row['keywords']
            original_question = row['original_question']
            
            logger.info(f"Processing FAQ item {idx + 1}/{total_items}: {original_question}")
            
            # 원본 질문 임베딩
            texts.append(original_question)
            metadatas.append({
                "type": "original_question",
                "structured_answer": json.dumps(structured_answer, ensure_ascii=False),
                "keywords": json.dumps(keywords, ensure_ascii=False)
            })
            ids.append(f"orig_{idx}")
            
            # 질문 변형들 임베딩
            for var_idx, q in enumerate(question_variations):
                texts.append(q)
                metadatas.append({
                    "type": "question_variation",
                    "original_question": original_question,
                    "structured_answer": json.dumps(structured_answer, ensure_ascii=False),
                    "keywords": json.dumps(keywords, ensure_ascii=False)
                })
                ids.append(f"q_{idx}_{var_idx}")
            
            # 구조화된 답변의 각 부분도 임베딩
            if 'basic_rules' in structured_answer and structured_answer['basic_rules']:
                for part_idx, rule in enumerate(structured_answer['basic_rules']):
                    text = f"기본 규칙: {rule}"
                    texts.append(text)
                    metadatas.append({
                        "type": "answer_part",
                        "part_type": "basic_rule",
                        "original_question": original_question,
                        "structured_answer": json.dumps(structured_answer, ensure_ascii=False),
                        "keywords": json.dumps(keywords, ensure_ascii=False)
                    })
                    ids.append(f"rule_{idx}_{part_idx}")
            
            if 'examples' in structured_answer and structured_answer['examples']:
                for part_idx, example in enumerate(structured_answer['examples']):
                    if isinstance(example, dict):
                        example_text = json.dumps(example, ensure_ascii=False)
                    else:
                        example_text = str(example)
                    
                    text = f"예시: {example_text}"
                    texts.append(text)
                    metadatas.append({
                        "type": "answer_part",
                        "part_type": "example",
                        "original_question": original_question,
                        "structured_answer": json.dumps(structured_answer, ensure_ascii=False),
                        "keywords": json.dumps(keywords, ensure_ascii=False)
                    })
                    ids.append(f"example_{idx}_{part_idx}")
            
            if 'cautions' in structured_answer and structured_answer['cautions']:
                for part_idx, caution in enumerate(structured_answer['cautions']):
                    text = f"주의사항: {caution}"
                    texts.append(text)
                    metadatas.append({
                        "type": "answer_part",
                        "part_type": "caution",
                        "original_question": original_question,
                        "structured_answer": json.dumps(structured_answer, ensure_ascii=False),
                        "keywords": json.dumps(keywords, ensure_ascii=False)
                    })
                    ids.append(f"caution_{idx}_{part_idx}")
        
        logger.info(f"생성할 임베딩 항목 수: {len(texts)}")
        
        # 임베딩 생성 및 저장
        if texts:
            logger.info("임베딩 생성 시작...")
            
            # 배치 크기 줄이기
            batch_size = 10  # 50에서 10으로 감소
            
            # embeddings 리스트 초기화
            embeddings = []
            
            # 메모리 사용량 최적화
            for i in range(0, len(texts), batch_size):
                batch_end = min(i + batch_size, len(texts))
                batch = texts[i:batch_end]
                
                logger.info(f"임베딩 생성 진행률: {batch_end / len(texts) * 100:.1f}% ({batch_end}/{len(texts)})")
                
                # 배치 단위로 임베딩 생성
                batch_embeddings = get_embeddings(batch)
                
                # 임베딩 형식 확인 및 수정
                if batch_embeddings:
                    # 만약 3차원 리스트라면 2차원으로 변환
                    while isinstance(batch_embeddings, list) and len(batch_embeddings) > 0 and isinstance(batch_embeddings[0], list) and isinstance(batch_embeddings[0][0], list):
                        batch_embeddings = [emb for sublist in batch_embeddings for emb in sublist]
                        logger.info("3차원 임베딩을 2차원으로 변환했습니다.")

                    # 만약 1차원 리스트(단일 벡터)라면 2차원으로 변환
                    if batch_embeddings and isinstance(batch_embeddings[0], (float, int)):
                        batch_embeddings = [batch_embeddings]
                        logger.info("단일 임베딩을 리스트로 변환했습니다.")

                    embeddings.extend(batch_embeddings)
                
                # 메모리 정리
                gc.collect()
                if torch.cuda.is_available():
                    torch.cuda.empty_cache()
                
                # 잠시 대기하여 메모리 해제 시간 확보
                await asyncio.sleep(0.1)
            
            logger.info("임베딩 생성 완료! 이제 ChromaDB에 저장합니다...")
            
            # 임베딩 형식 로깅 (디버깅용)
            if embeddings:
                logger.info(f"첫 번째 임베딩 차원 구조: {type(embeddings)}, {type(embeddings[0])}")
                if isinstance(embeddings[0], list):
                    logger.info(f"첫 번째 임베딩 길이: {len(embeddings[0])}")
            
            # 최종 임베딩 형식 검증
            if not all(isinstance(emb, list) and all(isinstance(x, (int, float)) for x in emb) for emb in embeddings):
                logger.error("임베딩 형식이 올바르지 않습니다.")
                return {"success": False, "message": "임베딩 형식이 올바르지 않습니다."}
            
            # === 임베딩 차원 강제 변환 (최종 방어) ===
            if isinstance(embeddings, np.ndarray):
                embeddings = embeddings.tolist()
            while isinstance(embeddings, list) and len(embeddings) > 0 and isinstance(embeddings[0], list) and isinstance(embeddings[0][0], list):
                embeddings = [emb for sublist in embeddings for emb in sublist]
            if embeddings and isinstance(embeddings[0], (float, int)):
                embeddings = [embeddings]
            # === 방어 끝 ===

            # 컬렉션에 추가
            collection.add(
                embeddings=embeddings,
                documents=texts,
                metadatas=metadatas,
                ids=ids
            )
            
            logger.info(f"벡터 데이터베이스에 {len(texts)}개 항목 추가 완료")
            
            # 간단한 테스트 쿼리 실행
            test_emb = get_embeddings("테스트 쿼리")
            while isinstance(test_emb, list) and len(test_emb) > 0 and isinstance(test_emb[0], list) and isinstance(test_emb[0][0], list):
                test_emb = [emb for sublist in test_emb for emb in sublist]
            if test_emb and isinstance(test_emb[0], (float, int)):
                test_emb = [test_emb]
            test_result = collection.query(
                query_embeddings=test_emb,
                n_results=1
            )
            logger.info(f"벡터 DB 테스트 쿼리 결과: {test_result}")
            
            return {"success": True, "message": f"{len(texts)}개 항목이 벡터 데이터베이스에 추가되었습니다."}
        else:
            logger.error("임베딩할 텍스트가 없습니다.")
            return {"success": False, "message": "임베딩할 텍스트가 없습니다."}
            
    except Exception as e:
        logger.error(f"벡터 데이터베이스 초기화 중 오류 발생: {str(e)}")
        return {"success": False, "message": f"오류: {str(e)}"}

# 파일 변경 감지를 위한 전역 변수
last_modified_time = None
file_check_interval = 5  # 5초마다 확인

class ExcelFileHandler(FileSystemEventHandler):
    def __init__(self, callback):
        self.callback = callback
        self.last_modified = 0
        
    def on_modified(self, event):
        if not event.is_directory and event.src_path.endswith('qa_pairs.xlsx'):
            # 중복 이벤트 방지를 위한 쿨다운 체크
            current_time = time.time()
            if current_time - self.last_modified > 1:  # 1초 쿨다운
                self.last_modified = current_time
                asyncio.create_task(self.callback())

async def check_file_changes():
    """qa_pairs.xlsx 파일의 변경을 주기적으로 확인합니다."""
    global last_modified_time
    
    while True:
        try:
            if ENHANCED_FAQ_PATH.exists():
                current_mtime = os.path.getmtime(ENHANCED_FAQ_PATH)
                if last_modified_time is None:
                    last_modified_time = current_mtime
                elif current_mtime > last_modified_time:
                    logger.info("enhanced_qa_pairs.xlsx 파일이 변경되었습니다. 벡터 DB를 업데이트합니다.")
                    last_modified_time = current_mtime
                    # 파일 변경 시에는 force_rebuild=True로 설정하여 임베딩 다시 생성
                    await initialize_vector_db(force_rebuild=True)
            await asyncio.sleep(file_check_interval)
        except Exception as e:
            logger.error(f"파일 변경 확인 중 오류 발생: {e}")
            await asyncio.sleep(file_check_interval)

# 마지막으로 정적 파일 서빙 설정하기 전에 추가
# Ping 엔드포인트 - 서버 활성 상태 유지용
@app.get("/ping")
async def ping():
    """서버가 활성 상태임을 확인하는 단순 엔드포인트"""
    return {"status": "alive", "timestamp": datetime.now().isoformat()}

# 시스템 정보 수집 함수
def collect_system_info():
    """시스템 리소스 사용량을 수집하여 반환합니다."""
    try:
        # 프로세스 정보 가져오기
        process = psutil.Process()
        
        # 메모리 정보
        memory_info = process.memory_info()
        memory_percent = process.memory_percent()
        
        # 시스템 전체 메모리 정보
        system_memory = psutil.virtual_memory()
        
        # CPU 사용량
        cpu_percent = process.cpu_percent(interval=1)
        
        # 디스크 사용량
        disk_usage = psutil.disk_usage('/')
        
        info = {
            "process_memory": {
                "rss": f"{memory_info.rss / 1024 / 1024:.2f} MB",  # 실제 물리적 메모리 사용량
                "vms": f"{memory_info.vms / 1024 / 1024:.2f} MB",  # 가상 메모리 사용량
                "percent": f"{memory_percent:.2f}%"
            },
            "system_memory": {
                "total": f"{system_memory.total / 1024 / 1024:.2f} MB",
                "available": f"{system_memory.available / 1024 / 1024:.2f} MB",
                "used": f"{system_memory.used / 1024 / 1024:.2f} MB",
                "percent": f"{system_memory.percent}%"
            },
            "cpu": {
                "percent": f"{cpu_percent}%"
            },
            "disk": {
                "total": f"{disk_usage.total / 1024 / 1024:.2f} MB",
                "used": f"{disk_usage.used / 1024 / 1024:.2f} MB",
                "free": f"{disk_usage.free / 1024 / 1024:.2f} MB",
                "percent": f"{disk_usage.percent}%"
            }
        }
        
        # 메모리 사용량이 80%를 초과하면 경고 로그
        if system_memory.percent > 80:
            logger.warning(f"메모리 사용량이 높습니다! ({system_memory.percent}%)")
        
        return info
    except Exception as e:
        logger.error(f"시스템 정보 수집 중 오류 발생: {e}")
        return {"error": str(e)}

@app.get("/system-info")
async def get_system_info():
    """시스템 리소스 사용량을 반환합니다."""
    return collect_system_info()

# 자동 핑을 위한 백그라운드 태스크
async def keep_alive():
    """서버가 슬립 상태로 전환되지 않도록 주기적으로 자체 핑을 수행하고 시스템 정보를 모니터링"""
    while True:
        try:
            # 5분마다 자체 ping (Render 무료 플랜 타임아웃은 일반적으로 15분)
            await asyncio.sleep(300)
            
            # 시스템 정보 수집 및 로깅
            system_info = collect_system_info()
            logger.info("시스템 리소스 사용량:")
            logger.info(f"- 프로세스 메모리 (RSS): {system_info['process_memory']['rss']}")
            logger.info(f"- 시스템 메모리 사용률: {system_info['system_memory']['percent']}")
            logger.info(f"- CPU 사용률: {system_info['cpu']['percent']}")
            
            async with httpx.AsyncClient() as client:
                # 현재 서버 URL 동적 생성
                host = "localhost"
                port = "8000"
                # 환경 변수 확인
                if os.getenv("RENDER") == "true":
                    # Render 환경에서는 외부 URL 사용
                    response = await client.get("https://construction-chatbot-api.onrender.com/ping")
                else:
                    # 개발 환경에서는 로컬 URL 사용
                    response = await client.get(f"http://{host}:{port}/ping")
                
                logger.info(f"자동 핑 응답: {response.status_code}")
        except Exception as e:
            logger.error(f"자동 핑 에러: {e}")
            continue

# 전역 변수로 태스크 저장
background_tasks = set()
active_processes = set()  # 활성 프로세스 추적을 위한 세트 추가

@app.on_event("startup")
async def startup_event():
    """서버 시작 시 실행되는 이벤트 핸들러"""
    logger.info("서버 시작: 벡터 데이터베이스 초기화 시작")
    
    # 초기 벡터 DB 초기화 (force_rebuild=False로 설정하여 기존 임베딩 재사용)
    init_result = await initialize_vector_db(force_rebuild=False)
    logger.info(f"벡터 DB 초기화 결과: {init_result}")
    
    # 파일 감시 시작
    global last_modified_time
    if ENHANCED_FAQ_PATH.exists():
        last_modified_time = os.path.getmtime(ENHANCED_FAQ_PATH)
    
    # 파일 변경 감지 태스크 시작
    file_check_task = asyncio.create_task(check_file_changes())
    background_tasks.add(file_check_task)
    file_check_task.add_done_callback(background_tasks.discard)
    
    # 서버 활성 상태 유지를 위한 자동 핑 태스크 시작
    keep_alive_task = asyncio.create_task(keep_alive())
    background_tasks.add(keep_alive_task)
    keep_alive_task.add_done_callback(background_tasks.discard)
    
    logger.info("파일 감시 및 자동 핑 시작됨")

@app.on_event("shutdown")
async def shutdown_event():
    """서버 종료 시 실행되는 이벤트 핸들러"""
    logger.info("서버 종료: 백그라운드 태스크 및 프로세스 정리 시작")
    
    # 모든 백그라운드 태스크 취소
    for task in background_tasks:
        task.cancel()
    
    # 태스크가 완료될 때까지 대기
    if background_tasks:
        await asyncio.gather(*background_tasks, return_exceptions=True)
    
    # 활성 프로세스 종료
    for process in active_processes:
        try:
            if process.poll() is None:  # 프로세스가 아직 실행 중인 경우
                process.terminate()  # 먼저 정상 종료 시도
                try:
                    process.wait(timeout=5)  # 5초 동안 대기
                except subprocess.TimeoutExpired:
                    process.kill()  # 강제 종료
        except Exception as e:
            logger.error(f"프로세스 종료 중 오류: {e}")
    
    # 메모리 정리
    gc.collect()
    if torch.cuda.is_available():
        torch.cuda.empty_cache()
    
    logger.info("서버 종료: 모든 백그라운드 태스크 및 프로세스 정리 완료")

# 수동 초기화용 엔드포인트 (필요시 재초기화 가능)
@app.get("/init-db")
async def manual_initialize_database():
    """벡터 데이터베이스를 수동으로 초기화합니다."""
    # 수동 초기화 시에는 force_rebuild=True로 설정
    success = await initialize_vector_db(force_rebuild=True)
    if success.get("success", False):
        return {"status": "success", "message": success.get("message", "데이터베이스가 성공적으로 초기화되었습니다.")}
    else:
        return {"status": "error", "message": success.get("message", "데이터베이스 초기화 중 오류가 발생했습니다.")}

@app.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    """Excel 파일을 업로드하고 저장합니다."""
    try:
        # 파일 확장자 확인
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Excel 파일(.xlsx 또는 .xls)만 업로드 가능합니다.")
        
        # 파일 저장 경로
        save_path = DOCS_DIR / "qa_pairs.xlsx"
        
        # 파일 저장
        content = await file.read()
        with open(save_path, "wb") as f:
            f.write(content)
        
        logger.info(f"Excel 파일이 성공적으로 업로드되었습니다: {save_path}")
        
        return JSONResponse(
            content={"message": "Excel 파일이 성공적으로 업로드되었습니다. '처리하기' 버튼을 눌러 데이터를 구조화해주세요."},
            status_code=200
        )
    except Exception as e:
        logger.error(f"Excel 파일 업로드 실패: {e}")
        raise HTTPException(status_code=500, detail=f"Excel 파일 업로드 실패: {str(e)}")

# process-excel 엔드포인트 수정
@app.post("/process-excel")
async def process_excel():
    """업로드된 FAQ 엑셀 파일을 처리하고 임베딩을 생성합니다."""
    try:
        qa_file_path = DOCS_DIR / "qa_pairs.xlsx"
        if not qa_file_path.exists():
            return JSONResponse(
                status_code=400,
                content={"success": False, "message": "처리할 엑셀 파일이 없습니다. 먼저 파일을 업로드해주세요."}
            )
        
        logger.info("FAQ 구조화 처리 시작...")
        
        # 비동기 작업으로 변경
        async def process_faq():
            try:
                script_path = Path(__file__).parent / "qa_allinone.py"
                
                if not script_path.exists():
                    raise FileNotFoundError(f"qa_allinone.py 파일을 찾을 수 없습니다: {script_path}")
                
                logger.info(f"qa_allinone.py 실행 시작: {script_path}")
                
                # 환경에 따른 subprocess 처리 방식 선택
                is_windows = os.name == 'nt'
                is_render = os.getenv("RENDER") == "true"
                
                if is_windows and not is_render:
                    # Windows 개발 환경
                    logger.info("Windows 개발 환경에서 실행")
                    process = subprocess.Popen(
                        [sys.executable, str(script_path)],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        text=True,
                        encoding='utf-8',
                        cwd=str(Path(__file__).parent),
                        env={**os.environ, 'PYTHONPATH': str(Path(__file__).parent)}
                    )
                    
                    # 프로세스를 활성 프로세스 세트에 추가
                    active_processes.add(process)
                    
                    # 실시간 로그 처리
                    error_output = []
                    
                    # stdout과 stderr를 비동기로 읽기
                    async def read_stream(stream, is_error=False):
                        while True:
                            line = stream.readline()
                            if not line and process.poll() is not None:
                                break
                            if line:
                                line = line.strip()
                                if is_error:
                                    error_output.append(line)
                                    logger.error(f"[qa_allinone.py] {line}")
                                else:
                                    logger.info(f"[qa_allinone.py] {line}")
                
                    # stdout과 stderr를 동시에 처리
                    await asyncio.gather(
                        read_stream(process.stdout),
                        read_stream(process.stderr, is_error=True)
                    )
                    
                    # 프로세스 종료 대기
                    return_code = process.wait()
                    
                    # 프로세스를 활성 프로세스 세트에서 제거
                    active_processes.discard(process)
                else:
                    # Render 배포 환경 또는 Linux/Mac
                    logger.info("Render 배포 환경 또는 Linux/Mac에서 실행")
                    process = await asyncio.create_subprocess_exec(
                        sys.executable, str(script_path),
                        stdout=asyncio.subprocess.PIPE,
                        stderr=asyncio.subprocess.PIPE,
                        cwd=str(Path(__file__).parent),
                        env={**os.environ, 'PYTHONPATH': str(Path(__file__).parent)}
                    )
                    
                    # 프로세스를 활성 프로세스 세트에 추가
                    active_processes.add(process)
                    
                    # 실시간 로그 처리
                    error_output = []
                    
                    # stdout과 stderr를 비동기로 읽기
                    async def read_stream(stream, is_error=False):
                        while True:
                            line = await stream.readline()
                            if not line and process.returncode is not None:
                                break
                            if line:
                                line = line.decode().strip()
                                if is_error:
                                    error_output.append(line)
                                    logger.error(f"[qa_allinone.py] {line}")
                                else:
                                    logger.info(f"[qa_allinone.py] {line}")
                    
                    # stdout과 stderr를 동시에 처리
                    await asyncio.gather(
                        read_stream(process.stdout),
                        read_stream(process.stderr, is_error=True)
                    )
                    
                    # 프로세스 종료 대기
                    return_code = await process.wait()
                    
                    # 프로세스를 활성 프로세스 세트에서 제거
                    active_processes.discard(process)
                
                logger.info(f"qa_allinone.py 종료 코드: {return_code}")
                
                if return_code != 0:
                    error_message = "\n".join(error_output) if error_output else "알 수 없는 오류"
                    raise Exception(f"qa_allinone.py 실행 실패 (종료 코드: {return_code}): {error_message}")
                
                # 벡터 DB 초기화
                logger.info("FAQ 구조화 처리 완료. 이제 새로운 구조화된 데이터로 벡터 데이터베이스를 초기화합니다.")
                
                if not ENHANCED_FAQ_PATH.exists():
                    raise FileNotFoundError(f"구조화된 FAQ 파일을 찾을 수 없습니다: {ENHANCED_FAQ_PATH}")
                
                if not load_enhanced_faq():
                    raise Exception("구조화된 FAQ 데이터를 로드할 수 없습니다.")
                
                db_result = await initialize_vector_db(force_rebuild=True)
                if not db_result["success"]:
                    raise Exception(f"벡터 데이터베이스 초기화 실패: {db_result['message']}")
                
                return True
                
            except Exception as e:
                logger.error(f"FAQ 처리 중 오류: {str(e)}")
                import traceback
                logger.error(f"상세 오류: {traceback.format_exc()}")
                return False
        
        # 비동기 작업 시작
        task = asyncio.create_task(process_faq())
        
        # 작업 완료 대기
        success = await task
        
        if success:
            return {"success": True, "message": "FAQ 처리 및 벡터 데이터베이스 초기화가 완료되었습니다."}
        else:
            return JSONResponse(
                status_code=500,
                content={"success": False, "message": "FAQ 처리 중 오류가 발생했습니다. 로그를 확인해주세요."}
            )
        
    except Exception as e:
        logger.error(f"FAQ 처리 시작 중 오류: {str(e)}")
        import traceback
        logger.error(f"상세 오류: {traceback.format_exc()}")
        return JSONResponse(
            status_code=500,
            content={"success": False, "message": f"FAQ 처리 시작 중 오류: {str(e)}"}
        )

# 정적 파일 서빙 설정은 모든 API 엔드포인트 정의 후 맨 마지막에 위치
if FRONTEND_BUILD_DIR.exists():
    app.mount("/", StaticFiles(directory=str(FRONTEND_BUILD_DIR), html=True), name="frontend")
    logger.info(f"Frontend static files mounted from {FRONTEND_BUILD_DIR}")
else:
    logger.warning(f"Frontend build directory not found at {FRONTEND_BUILD_DIR}. Static file serving is disabled.")

@app.get("/download-questions")
async def download_questions_excel():
    """저장된 질문을 Excel 파일로 다운로드합니다."""
    try:
        # 임시 파일 생성
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_path = temp_file.name
        
        # 기존 CSV 파일이 있는지 확인
        if not QUESTIONS_FILE.exists():
            # 파일이 없으면 빈 데이터프레임 생성
            df = pd.DataFrame(columns=['timestamp', 'session_id', 'query', 'answer', 'similarity'])
        else:
            # CSV 파일 읽기
            try:
                df = pd.read_csv(QUESTIONS_FILE, encoding='utf-8-sig')
            except Exception as e:
                logger.error(f"CSV 파일 읽기 오류: {e}")
                df = pd.DataFrame(columns=['timestamp', 'session_id', 'query', 'answer', 'similarity'])
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "질문 데이터"
        
        # 헤더 추가
        headers = ['시간', '세션 ID', '질문', '답변', '유사도']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            # 헤더 스타일 설정
            cell.font = cell.font.copy(bold=True)
        
        # 데이터 추가
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx+2, column=col_idx, value=value)
        
        # 열 너비 자동 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            # 최대 너비 제한 (너무 넓지 않게)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # 파일 저장
        wb.save(temp_path)
        
        # 파일 다운로드 응답
        return FileResponse(
            path=temp_path,
            filename="saved_questions.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Excel 파일 생성 중 오류 발생: {e}")
        import traceback
        logger.error(f"상세 오류: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/download-enhanced-qa")
async def download_enhanced_qa(filename: str = Query("enhanced_qa_pairs.xlsx")):
    """구조화된 enhanced_qa_pairs.xlsx 파일을 다운로드합니다."""
    try:
        if not ENHANCED_FAQ_PATH.exists():
            raise HTTPException(status_code=404, detail="enhanced_qa_pairs.xlsx 파일이 존재하지 않습니다.")
        return FileResponse(
            path=ENHANCED_FAQ_PATH,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"enhanced_qa_pairs.xlsx 다운로드 중 오류: {e}")
        raise HTTPException(status_code=500, detail=str(e))